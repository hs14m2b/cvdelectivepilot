import csv
import logging
import pathlib
import re
import time
import urllib
import os
from openpyxl import load_workbook
#import phonenumbers
#from phonenumbers import NumberParseException
from io import BytesIO

import boto3

# Set up logging
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

logger.debug('This will get logged')

# Config Items
SOURCE_XLSX_FILENAME = "Trust Submission Template - v2.xlsx"
SOURCE_FOLDER = "upload"

RESULTS_FOLDER = os.getenv(key="RESULTS_FOLDER")
OUTPUT_FOLDER = os.getenv(key="OUTPUT_FOLDER")
S3BUCKETNAME = os.getenv(key="S3BUCKETNAME")

# Set up S3 client
s3 = boto3.client('s3')

def lambda_handler(event, context):

    logger.debug(f"event is {event}")
    bucket = event['Records'][0]['s3']['bucket']['name']
    logger.debug(f"bucket is {bucket}")
    key = urllib.parse.unquote_plus(event['Records'][0]['s3']['object']['key'], encoding='utf-8')
    logger.debug(f"key is {key}")
    file_name = key.split('/')[1].replace(' ', '_')
    logger.debug(f"file_name is {file_name}")
    file_name_prefix = file_name.split('.')[0]
    logger.debug(f"file_name_prefix is {file_name_prefix}")
    OUTPUT_CSV_FILENAME = file_name_prefix+'.csv'
    logger.debug(f"OUTPUT_CSV_FILENAME is {OUTPUT_CSV_FILENAME}")
    RESULTS_CSV_FILENAME = file_name_prefix+'_results.csv'
    logger.debug(f"RESULTS_CSV_FILENAME is {RESULTS_CSV_FILENAME}")

    obj = s3.get_object(Bucket=bucket, Key=key) 
    binary_data = obj['Body'].read()
    
    # Load workbook
    # workbook = load_workbook(filename=SOURCE_XLSX_FILENAME, read_only=True)
    workbook = load_workbook(BytesIO(binary_data))

    # Get message content
    config_sheet = workbook["Config"]
    message = config_sheet["B1"].value

    # Get list of mobile numbers
    sheet = workbook["Patient Details"]
    data = sheet["C3:C100"]

    export_list = []

    processed_success = 0
    processed_error = 0

    processed_results = []

    for cell in data:
        number = str(cell[0].value)
        if number and number != "None":
            logger.debug(f"number is {number}")
            try:
                #phonenumbers.parse(number, "GB")
                regex_result = re.search("^[0+]\d{10,14}$", number)
                if not regex_result:
                    raise ValueError("Does not match regex")
                export_list.append(cell)
                logger.info(number)
                processed_success += 1
                processed_results.append((number, "Processed"))

            except (ValueError) as e:
                logger.error(e)
                processed_error += 1
                processed_results.append((number, e))


    # Write numbers out to CSV
    with open('/tmp/'+OUTPUT_CSV_FILENAME, mode='w') as number_list_file:
        writer = csv.writer(number_list_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)

        row_count = 0
        logger.debug("Writing header to csv")
        writer.writerow(['mobile', 'message'])

        for row in export_list:
            #print(row)
            number = str(row[0].value)
            writer.writerow([number, message])
            row_count += 1

        print(row_count)    
        logger.debug(f"Wrote {row_count} rows to {OUTPUT_CSV_FILENAME}")

    # Move file to processed folder

    #source = pathlib.Path(SOURCE_XLSX_FILENAME)
    #processed_folder = pathlib.Path(OUTPUT_FOLDER)

    #destination = processed_folder.joinpath(SOURCE_XLSX_FILENAME)
    #print(destination)

    #if not destination.exists():
    #    source.replace(destination)

    # Write results out to CSV
    with open('/tmp/'+RESULTS_CSV_FILENAME, mode='w') as results_list_file:
        writer = csv.writer(results_list_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)

        row_count = 0

        for item in processed_results:
            writer.writerow([item[0], item[1]])
            row_count += 1

        print(row_count)    
        logger.debug(f"Wrote {row_count} rows to {RESULTS_CSV_FILENAME}")

    #copy csv to S3 inbox
    s3.upload_file('/tmp/'+OUTPUT_CSV_FILENAME, S3BUCKETNAME, OUTPUT_FOLDER+'/'+OUTPUT_CSV_FILENAME)
    logger.debug(f"Copied output file to {OUTPUT_FOLDER} folder")

    s3.upload_file('/tmp/'+RESULTS_CSV_FILENAME, S3BUCKETNAME, RESULTS_FOLDER+'/'+RESULTS_CSV_FILENAME)
    logger.debug(f"Copied results file to {RESULTS_FOLDER} folder")

    s3.delete_object(Bucket=bucket, Key=key)
    logger.debug(f"Deleted original inbound file {key}")

    logger.info(f"Results: {processed_success} number(s) processed succesfully, {processed_error} number(s) failed.")
